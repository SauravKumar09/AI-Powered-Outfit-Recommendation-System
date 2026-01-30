"""
Import products from a CSV/Excel URL (Google Sheets export) or local file.

Usage:
    python manage.py import_products --url <csv_or_xlsx_url>
    python manage.py import_products --file <path_to_file>

Defaults to the provided Google Sheet if no URL is given.
"""

import csv
import io
import sys
from urllib.request import urlopen
from urllib.parse import urlencode

from django.core.management.base import BaseCommand, CommandError

from apps.products.utils import import_products_from_workbook_rows


DEFAULT_SHEET_ID = "1bSdUJsST5sgi2brk1AFDKz0TS9zafaMT2_DJQHFsOWI"
DEFAULT_SHEET_EXPORT = (
    f"https://docs.google.com/spreadsheets/d/{DEFAULT_SHEET_ID}/export?format=csv"
)


class Command(BaseCommand):
    help = "Import products from a CSV or Excel source (Google Sheets supported)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--url", dest="url", help="CSV/XLSX URL. Defaults to provided Google Sheet."
        )
        parser.add_argument("--file", dest="file", help="Local CSV/XLSX file path.")

    def handle(self, *args, **options):
        source_url = options.get("url") or DEFAULT_SHEET_EXPORT
        file_path = options.get("file")

        if source_url and file_path:
            raise CommandError("Provide either --url or --file, not both.")

        if file_path:
            self.stdout.write(self.style.NOTICE(f"Importing from file {file_path}"))
            with open(file_path, "rb") as fh:
                content = fh.read()
        else:
            self.stdout.write(self.style.NOTICE(f"Fetching {source_url}"))
            try:
                with urlopen(source_url) as resp:
                    content = resp.read()
            except Exception as exc:
                raise CommandError(f"Failed to fetch URL: {exc}")

        # Decide by magic bytes / extension
        if file_path:
            lower_name = file_path.lower()
        else:
            lower_name = source_url.lower()

        is_csv = lower_name.endswith(".csv")

        try:
            if is_csv:
                decoded = content.decode("utf-8")
                reader = csv.reader(io.StringIO(decoded))
                rows = list(reader)
                if not rows:
                    raise CommandError("Empty CSV content")
                headers = [
                    str(h).strip().lower() if h is not None else "" for h in rows[0]
                ]
                result = import_products_from_workbook_rows(rows[1:], headers)
            else:
                import openpyxl  # Lazy import to avoid cost when not needed

                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    raise CommandError("Empty spreadsheet")
                headers = [
                    str(h).strip().lower() if h is not None else "" for h in rows[0]
                ]
                result = import_products_from_workbook_rows(rows[1:], headers)
        except CommandError:
            raise
        except Exception as exc:
            raise CommandError(f"Import failed: {exc}")

        created = result.get("created", 0)
        errors = result.get("errors", [])
        self.stdout.write(self.style.SUCCESS(f"Imported {created} products"))
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} rows had errors"))
            for err in errors[:5]:
                self.stdout.write(f"Row {err.get('row')}: {err.get('errors')}")
            if len(errors) > 5:
                self.stdout.write("... (truncated) ...")

        return 0
