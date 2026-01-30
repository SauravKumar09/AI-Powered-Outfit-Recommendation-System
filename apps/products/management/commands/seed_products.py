"""
Management command to seed the database with products from Sample_Products.xlsx.
"""

import ast
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.products.models import Product, ProductOccasion, ProductSeason


class Command(BaseCommand):
    help = "Seed the database with products from Sample_Products.xlsx"

    FILE_NAME = "Sample_Products.xlsx"

    COLOR_KEYWORDS = [
        "black",
        "white",
        "gray",
        "grey",
        "navy",
        "blue",
        "light blue",
        "sky blue",
        "red",
        "maroon",
        "green",
        "olive",
        "beige",
        "khaki",
        "tan",
        "brown",
        "purple",
        "burgundy",
        "yellow",
        "gold",
        "silver",
        "pink",
    ]

    def handle(self, *args, **options):
        base_dir = Path(__file__).resolve().parents[4]
        excel_path = base_dir / self.FILE_NAME
        if not excel_path.exists():
            raise CommandError(f"Seed file not found: {excel_path}")

        self.stdout.write(f"Loading data from {excel_path}...")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise CommandError("openpyxl is required to run this command") from exc

        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        products_data = []
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(value is None for value in row):
                continue

            data = dict(zip(headers, row))
            name = (data.get("title") or data.get("name") or "").strip()
            if not name:
                skipped += 1
                continue

            category = self.map_category(data.get("sub_category"), data.get("product_type"))
            if not category:
                skipped += 1
                continue

            style = self.map_style(data.get("product_type"), data.get("sub_category"), name, data.get("tags"))
            gender = self.map_gender(data.get("gender"))
            price = self.normalize_price(data.get("lowest_price"))
            price_range = self.map_price_range(price)
            color = self.extract_color(name, data.get("tags"), data.get("sub_category"))
            tags = self.parse_tags(data.get("tags"))
            image_url = data.get("featured_image") or ""
            description = data.get("description") or ""
            sku = (data.get("sku_id") or "").strip() or None

            occasions = self.derive_occasions(style)
            seasons = self.derive_seasons(name, data.get("sub_category"))

            products_data.append(
                {
                    "name": name,
                    "category": category,
                    "sub_category": data.get("sub_category") or (data.get("product_type") or "misc"),
                    "color": color,
                    "style": style,
                    "gender": gender,
                    "price": price,
                    "price_range": price_range,
                    "image_url": image_url,
                    "tags": tags,
                    "description": description,
                    "sku": sku,
                    "occasions": occasions,
                    "seasons": seasons,
                }
            )

        if not products_data:
            raise CommandError("No products to import from spreadsheet")

        self.stdout.write(f"Importing {len(products_data)} products (skipped {skipped})...")

        with transaction.atomic():
            ProductOccasion.objects.all().delete()
            ProductSeason.objects.all().delete()
            Product.objects.all().delete()

            occasion_relations = []
            season_relations = []

            for product_data in products_data:
                occasions = product_data.pop("occasions", [])
                seasons = product_data.pop("seasons", [])

                product = Product.objects.create(**product_data)

                for occasion in occasions:
                    occasion_relations.append(ProductOccasion(product=product, occasion=occasion))

                for season in seasons:
                    season_relations.append(ProductSeason(product=product, season=season))

            if occasion_relations:
                ProductOccasion.objects.bulk_create(occasion_relations, ignore_conflicts=True)
            if season_relations:
                ProductSeason.objects.bulk_create(season_relations, ignore_conflicts=True)

        self.stdout.write(self.style.SUCCESS(f"Seeded {len(products_data)} products (skipped {skipped})"))

    def parse_tags(self, raw):
        if isinstance(raw, list):
            return [str(tag).strip() for tag in raw if tag]
        if isinstance(raw, str):
            try:
                parsed = ast.literal_eval(raw)
                if isinstance(parsed, list):
                    return [str(tag).strip() for tag in parsed if tag]
            except (SyntaxError, ValueError):
                pass
            return [part.strip() for part in raw.split(",") if part.strip()]
        return []

    def map_category(self, sub_category, product_type):
        text = f"{sub_category or ''} {product_type or ''}".lower()
        top_keys = ["shirt", "tee", "tshirt", "t-shirt", "polo", "hoodie", "sweatshirt", "jacket", "coat", "blazer", "sweater", "cardigan", "kurta", "top"]
        bottom_keys = ["jean", "trouser", "pant", "chino", "short", "cargo", "jogger", "skirt", "bottom"]
        footwear_keys = ["shoe", "sneaker", "boot", "loafer", "sandal", "slipper", "flip flop", "heel"]
        accessory_keys = ["belt", "wallet", "bag", "backpack", "watch", "sunglass", "glasses", "cap", "hat", "beanie", "scarf", "tie", "pocket square", "bracelet"]

        if any(key in text for key in top_keys):
            return "top"
        if any(key in text for key in bottom_keys):
            return "bottom"
        if any(key in text for key in footwear_keys):
            return "footwear"
        if any(key in text for key in accessory_keys):
            return "accessory"
        return None

    def map_style(self, product_type, sub_category, name, tags):
        text = " ".join(str(part or "") for part in [product_type, sub_category, name, tags]).lower()
        if any(key in text for key in ["oxford", "trouser", "formal", "blazer", "tie", "belt"]):
            return "formal"
        if any(key in text for key in ["chino", "polo", "loafer", "chelsea", "desert", "smart"]):
            return "smart_casual"
        if any(key in text for key in ["run", "sport", "athletic", "jogger", "sneaker"]):
            return "sporty"
        return "casual"

    def map_gender(self, value):
        if not value:
            return "unisex"
        text = str(value).lower()
        if "female" in text or text == "women" or text == "woman":
            return "female"
        if "male" in text or text == "men" or text == "man":
            return "male"
        return "unisex"

    def normalize_price(self, value):
        try:
            price = Decimal(str(value))
        except Exception:  # pragma: no cover
            price = Decimal("0")
        if price < 0:
            price = Decimal("0")
        return price.quantize(Decimal("0.01"))

    def map_price_range(self, price):
        if price < Decimal("50"):
            return "budget"
        if price < Decimal("150"):
            return "mid"
        if price < Decimal("300"):
            return "premium"
        return "luxury"

    def extract_color(self, name, tags, sub_category):
        text = " ".join(str(part or "") for part in [name, tags, sub_category]).lower()
        for color in self.COLOR_KEYWORDS:
            if color in text:
                return color.replace(" ", "_")
        return "multi"

    def derive_occasions(self, style):
        if style == "formal":
            return ["office", "interview", "wedding", "formal"]
        if style == "smart_casual":
            return ["office", "casual", "date"]
        if style == "sporty":
            return ["casual", "weekend", "outdoor"]
        return ["casual", "weekend"]

    def derive_seasons(self, name, sub_category):
        text = " ".join(str(part or "") for part in [name, sub_category]).lower()
        if any(key in text for key in ["hoodie", "sweater", "jacket", "coat", "wool", "puffer"]):
            return ["fall", "winter", "spring"]
        if any(key in text for key in ["short", "linen", "tee", "t-shirt", "tank"]):
            return ["spring", "summer"]
        return ["all"]