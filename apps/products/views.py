"""
Product views.
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter

from .models import Product
from .serializers import (
    ProductSerializer,
    ProductListSerializer,
    ProductCreateSerializer,
)
import openpyxl
from io import BytesIO
import csv
import io
from .utils import import_products_from_workbook_rows


@extend_schema_view(
    list=extend_schema(
        tags=["Products"],
        summary="List all products",
        description="Get a list of all products with optional filtering.",
        parameters=[
            OpenApiParameter(
                name="category",
                description="Filter by category (top, bottom, footwear, accessory)",
            ),
            OpenApiParameter(
                name="style",
                description="Filter by style (formal, smart_casual, casual, sporty)",
            ),
            OpenApiParameter(name="color", description="Filter by color"),
            OpenApiParameter(name="price_range", description="Filter by price range"),
        ],
    ),
    retrieve=extend_schema(
        tags=["Products"],
        summary="Get product details",
        description="Get detailed information about a specific product.",
    ),
    create=extend_schema(
        tags=["Products"],
        summary="Create a new product",
        description="Add a new product to the catalog.",
    ),
    update=extend_schema(
        tags=["Products"],
        summary="Update a product",
        description="Update all fields of an existing product.",
    ),
    partial_update=extend_schema(
        tags=["Products"],
        summary="Partially update a product",
        description="Update specific fields of an existing product.",
    ),
    destroy=extend_schema(
        tags=["Products"],
        summary="Delete a product",
        description="Remove a product from the catalog.",
    ),
)
class ProductViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Product CRUD operations.
    """

    queryset = Product.objects.filter(is_active=True).prefetch_related(
        "occasions", "seasons"
    )
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["category", "style", "color", "price_range", "gender"]
    search_fields = ["name", "sub_category", "tags"]
    ordering_fields = ["price", "name", "created_at"]
    ordering = ["category", "name"]

    def get_serializer_class(self):
        if self.action == "list":
            return ProductListSerializer
        elif self.action == "create":
            return ProductCreateSerializer
        return ProductSerializer

    @extend_schema(
        tags=["Products"],
        summary="Get products by category",
        description="Get all products in a specific category.",
        parameters=[
            OpenApiParameter(
                name="category",
                description="Category name",
                required=True,
                type=str,
                location=OpenApiParameter.PATH,
            )
        ],
    )
    @action(detail=False, methods=["get"], url_path="category/(?P<category>[^/.]+)")
    def by_category(self, request, category=None):
        """
        Get products by category.
        """
        products = self.queryset.filter(category=category)
        serializer = ProductListSerializer(products, many=True)
        return Response(
            {
                "success": True,
                "category": category,
                "count": products.count(),
                "products": serializer.data,
            }
        )

    @extend_schema(
        tags=["Products"],
        summary="Get available filters",
        description="Get all available filter options for products.",
    )
    @action(detail=False, methods=["get"])
    def filters(self, request):
        """
        Get available filter options.
        """
        return Response(
            {
                "success": True,
                "filters": {
                    "categories": list(
                        Product.objects.values_list("category", flat=True).distinct()
                    ),
                    "styles": list(
                        Product.objects.values_list("style", flat=True).distinct()
                    ),
                    "colors": list(
                        Product.objects.values_list("color", flat=True).distinct()
                    ),
                    "price_ranges": list(
                        Product.objects.values_list("price_range", flat=True).distinct()
                    ),
                    "genders": list(
                        Product.objects.values_list("gender", flat=True).distinct()
                    ),
                },
            }
        )

    @extend_schema(
        tags=["Products"],
        summary="Upload products spreadsheet",
        description="Accepts an Excel (.xlsx) or CSV file and imports products in bulk.",
    )
    @action(detail=False, methods=["post"], url_path="upload")
    def upload(self, request):
        """
        Upload a spreadsheet (`.xlsx` or `.csv`) with product rows. First row must be headers.
        Supported header names (case-insensitive): name, category, sub_category, color,
        image_url/featured_image, style, gender, price/lowest_price, price_range, tags,
        occasions, seasons, sku. `occasions` and `seasons` may be comma-separated lists in a single cell.
        """
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response(
                {"success": False, "error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            name_lower = file_obj.name.lower()
            if name_lower.endswith(".csv"):
                decoded = file_obj.read().decode("utf-8")
                reader = csv.reader(io.StringIO(decoded))
                rows = list(reader)
                if not rows:
                    return Response(
                        {"success": False, "error": "Empty spreadsheet"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                headers = [
                    str(h).strip().lower() if h is not None else "" for h in rows[0]
                ]
                result = import_products_from_workbook_rows(rows[1:], headers)
            else:
                wb = openpyxl.load_workbook(
                    filename=BytesIO(file_obj.read()), read_only=True
                )
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return Response(
                        {"success": False, "error": "Empty spreadsheet"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                headers = [
                    str(h).strip().lower() if h is not None else "" for h in rows[0]
                ]
                result = import_products_from_workbook_rows(rows[1:], headers)
            return Response(
                {
                    "success": True,
                    "created": result.get("created", 0),
                    "errors": result.get("errors", []),
                }
            )

        except Exception as exc:
            return Response(
                {"success": False, "error": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
