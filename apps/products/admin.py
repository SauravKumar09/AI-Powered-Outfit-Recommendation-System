"""
Product admin configuration.
"""

from django.contrib import admin
from .models import Product, ProductOccasion, ProductSeason
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from io import BytesIO
import openpyxl
from .utils import import_products_from_workbook_rows


class ProductOccasionInline(admin.TabularInline):
    model = ProductOccasion
    extra = 1


class ProductSeasonInline(admin.TabularInline):
    model = ProductSeason
    extra = 1


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = [
        "sku",
        "name",
        "category",
        "sub_category",
        "color",
        "style",
        "price",
        "price_range",
        "is_active",
    ]
    list_filter = ["category", "style", "price_range", "gender", "is_active"]
    search_fields = ["sku", "name", "sub_category", "color"]
    ordering = ["category", "name"]
    inlines = [ProductOccasionInline, ProductSeasonInline]

    fieldsets = (
        (
            "Basic Information",
            {
                "fields": (
                    "sku",
                    "name",
                    "category",
                    "sub_category",
                    "color",
                    "image_url",
                    "description",
                )
            },
        ),
        ("Classification", {"fields": ("style", "gender", "tags")}),
        ("Pricing", {"fields": ("price", "price_range")}),
        ("Status", {"fields": ("is_active",)}),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "upload-excel/",
                self.admin_site.admin_view(self.upload_excel),
                name="products_upload_excel",
            )
        ]
        return custom + urls

    def upload_excel(self, request):
        if request.method == "POST":
            f = request.FILES.get("file")
            if not f:
                messages.error(request, "No file uploaded")
                return redirect("..")
            try:
                wb = openpyxl.load_workbook(filename=BytesIO(f.read()), read_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    messages.error(request, "Empty spreadsheet")
                    return redirect("..")
                headers = [
                    str(h).strip().lower() if h is not None else "" for h in rows[0]
                ]
                result = import_products_from_workbook_rows(rows[1:], headers)
                messages.success(
                    request,
                    f"Imported {result.get('created', 0)} rows; {len(result.get('errors', []))} errors",
                )
            except Exception as exc:
                messages.error(request, f"Import failed: {exc}")
            return redirect("..")

        # GET: render simple upload form
        return render(request, "admin/products/upload.html", {})


# Also register the Occasion and Season models individually for quick access
@admin.register(ProductOccasion)
class ProductOccasionAdmin(admin.ModelAdmin):
    list_display = ["product", "occasion"]
    search_fields = ["product__name", "occasion"]


@admin.register(ProductSeason)
class ProductSeasonAdmin(admin.ModelAdmin):
    list_display = ["product", "season"]
    search_fields = ["product__name", "season"]
